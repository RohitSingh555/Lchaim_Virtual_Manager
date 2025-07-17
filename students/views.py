from datetime import date, datetime, time, timedelta
import decimal
import json
import os
import re
import threading
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
# Django Imports
from django.shortcuts import render, get_object_or_404, redirect
from .models import StudentProfile, VolunteerLog, ActivityLog
from datetime import timedelta
import re
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db.models import Case, When, Count, Min, Q, Sum, Value, IntegerField
from django.http import HttpResponse, FileResponse, Http404, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.core.files.storage import default_storage
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from io import BytesIO
from openpyxl import Workbook
from .forms import CollegeForm, OrientationDateForm, StudentFileForm, StudentProfileForm, VolunteerLogForm
from .models import College, OrientationDate, Shift, StudentFile, StudentProfile, VolunteerLog
from .utils import log_activity
from django.views.decorators.http import require_POST


def admin_required(function):
    def wrap(request, *args, **kwargs):
        if not request.user.groups.filter(name='Admin').exists():
            return redirect('student_profile_list') 
        return function(request, *args, **kwargs)
    return wrap

def school_required(function):
    def wrap(request, *args, **kwargs):
        if not request.user.groups.filter(name='School').exists():
            return redirect('student_profile_list') 
        return function(request, *args, **kwargs)
    return wrap

@admin_required
def signup_view(request):
    if request.user.is_authenticated:
        return redirect('student_profile_list')  
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Signup successful! You can now log in.')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'authentication/signup.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('student_profile_list')  
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('student_profile_list')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    return render(request, 'authentication/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

from datetime import datetime, date, timedelta



def calculate_shift_total_time(start_time, end_time):
    dt_start = datetime.combine(date.min, start_time)
    dt_end = datetime.combine(date.min, end_time)
    if dt_end < dt_start:
        # Overnight shift: add one day to end time
        dt_end += timedelta(days=1)
    duration = dt_end - dt_start
    hours = abs(duration.total_seconds()) / 3600
    return hours

# def get_shift_availability(user, start_date, shift_type, requested_hours):
#     try:
#         print('Shift type:', shift_type)    
#         assigned_shift = Shift.objects.get(id=shift_type)
#     except Shift.DoesNotExist:
#         return {"error": f"Shift type '{shift_type}' does not exist."}

#     # shift_total_time = (
#     #     datetime.combine(date.min, assigned_shift.end_time)
#     #     - datetime.combine(date.min, assigned_shift.start_time)
#     # ).seconds / 3600
#     shift_total_time = calculate_shift_total_time(assigned_shift.start_time,assigned_shift.end_time)

#     days_required = (requested_hours // shift_total_time) + (
#         1 if requested_hours % shift_total_time != 0 else 0
#     )

#     current_date = start_date
#     working_days = 0
#     shift_type = assigned_shift.type.lower()
#     if "weekendday" in shift_type:
#         weekdays_selected = {
#             "Saturday": 5,
#             "Sunday": 6
#         }
#     elif "weekendnight" in shift_type:
#         weekdays_selected = {
#             "Friday": 4,
#             "Saturday": 5,
#             "Sunday": 6
#         }
#         # weekdays_selected = [4, 5, 6]
#     elif "night" in shift_type:
#         weekdays_selected = {
#             "Monday": 0,
#             "Tuesday": 1,
#             "Wednesday": 2,
#             "Thursday": 3,
#         }
#         # weekdays_selected = [0, 1, 2, 3]
#     else:
#         weekdays_selected = {
#             "Monday": 0,
#             "Tuesday": 1,
#             "Wednesday": 2,
#             "Thursday": 3,
#             "Friday": 4
#         }
        
#     selected_weekdays = set(weekdays_selected.values())

#     while working_days < days_required:
#         if current_date.weekday() in selected_weekdays:
#             if validate_shift_capacity(user, current_date, assigned_shift):
#                 working_days += 1
#             else:
#                 return {
#                     "is_available": False,
#                     "message": f"Capacity exceeded on {current_date.strftime('%Y-%m-%d')} for {shift_type} shift.",
#                 }

#         current_date += timedelta(days=1)

#     end_date = current_date - timedelta(days=1)

#     return {
#         "is_available": True,
#         "start_date": start_date,
#         "shift_type": shift_type,
#         "requested_hours": requested_hours,
#         # "shift_requested": shift_requested,
#         "weekdays_selected": list(selected_weekdays),
#         "end_date": end_date,
#     }


def get_shift_availability(user, start_date, shift_type, requested_hours):
    try:
        print('Shift type:', shift_type)    
        assigned_shift = Shift.objects.get(id=shift_type)
    except Shift.DoesNotExist:
        return {"error": f"Shift type '{shift_type}' does not exist."}

    shift_total_time = calculate_shift_total_time(assigned_shift.start_time, assigned_shift.end_time)
    days_required = (requested_hours // shift_total_time) + (
        1 if requested_hours % shift_total_time != 0 else 0
    )

    shift_type_lower = assigned_shift.type.lower()
    
    # Define weekdays based on shift type
    if "weekendday" in shift_type_lower:
        weekdays_selected = {"Saturday": 5, "Sunday": 6}
    elif "weekendnight" in shift_type_lower:
        weekdays_selected = {"Friday": 4, "Saturday": 5, "Sunday": 6}
    elif "night" in shift_type_lower:
        weekdays_selected = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3}
    else:
        weekdays_selected = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3, "Friday": 4}
        
    selected_weekdays = set(weekdays_selected.values())

    # First, check if the requested start date works
    temp_date = start_date
    temp_working_days = 0
    capacity_exceeded_date = None
    
    while temp_working_days < days_required:
        if temp_date.weekday() in selected_weekdays:
            if validate_shift_capacity(user, temp_date, assigned_shift):
                temp_working_days += 1
            else:
                capacity_exceeded_date = temp_date
                break
        temp_date += timedelta(days=1)

    # If original start date works, return success
    if capacity_exceeded_date is None:
        end_date = temp_date - timedelta(days=1)
        return {
            "is_available": True,
            "start_date": start_date.strftime('%Y-%m-%d'),
            "shift_type": assigned_shift.type,
            "requested_hours": requested_hours,
            "weekdays_selected": list(selected_weekdays),
            "end_date": end_date.strftime('%Y-%m-%d'),
        }
    
    # If original start date doesn't work, find next available date
    next_available_date = capacity_exceeded_date + timedelta(days=1)
    max_search_days = 365  # Limit search to avoid infinite loops
    search_days = 0
    
    while search_days < max_search_days:
        # Try starting from next_available_date
        temp_date = next_available_date
        temp_working_days = 0
        can_complete_schedule = True
        
        while temp_working_days < days_required:
            if temp_date.weekday() in selected_weekdays:
                if validate_shift_capacity(user, temp_date, assigned_shift):
                    temp_working_days += 1
                else:
                    can_complete_schedule = False
                    break
            temp_date += timedelta(days=1)
        
        if can_complete_schedule:
            # Found a valid start date
            end_date = temp_date - timedelta(days=1)
            return {
                "is_available": False,
                "message": f"Capacity exceeded on {capacity_exceeded_date.strftime('%Y-%m-%d')} for {assigned_shift.type} shift.",
                "next_available_start_date": next_available_date.strftime('%Y-%m-%d'),
                "suggested_end_date": end_date.strftime('%Y-%m-%d'),
                "shift_type": assigned_shift.type,
                "requested_hours": requested_hours,
                "weekdays_selected": list(selected_weekdays),
            }
        
        # Move to next day and continue searching
        next_available_date += timedelta(days=1)
        search_days += 1
    
    # If no available date found within search limit
    return {
        "is_available": False,
        "message": f"Capacity exceeded on {capacity_exceeded_date.strftime('%Y-%m-%d')} for {assigned_shift.type} shift. No available slots found in the next {max_search_days} days.",
        "shift_type": assigned_shift.type,
        "requested_hours": requested_hours,
        "weekdays_selected": list(selected_weekdays),
    }



@csrf_exempt 
def shift_availability_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            
            start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
            shift_type = data['shift_type']
            requested_hours = int(data['requested_hours'])

            
            availability = get_shift_availability(request.user, start_date, shift_type, requested_hours)
            return JsonResponse(availability)
        
        except KeyError as e:
            return JsonResponse({"error": f"Missing key: {str(e)}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method."}, status=405)

def validate_shift_capacity(user, date, assigned_shift):
    """
    Validates if the shift on the given date exceeds its capacity.
    Superusers can always be assigned shifts regardless of capacity.
    """
    if user.is_superuser:
        return True

    shift_students_count = VolunteerLog.objects.filter(date=date, shift=assigned_shift).count()
    return shift_students_count < assigned_shift.max_students

@login_required
def create_student_profile(request):
    shifts = Shift.objects.all()  
    if request.method == "POST":
        profile_form = StudentProfileForm(request.POST, request.FILES)
        if profile_form.is_valid():
            student_profile = profile_form.save()
            file_fields = [
                "covid_vaccination",
                "medical_certificate",
                "police_check",
                "cpr_first_aid",
                "mask_fit"
            ]

            for field_name in file_fields:
                file = request.FILES.get(field_name)
                if file:
                    StudentFile.objects.create(student=student_profile, file=file)

            shift_id = request.POST.get('shift_timing')
            print("Shift ID:", shift_id)
            college_request_id = request.POST.get('college')
            college = College.objects.get(id=college_request_id)
            student_profile.school = college.name

            try:
                assigned_shift = Shift.objects.get(id=shift_id)
                student_profile.assigned_shift = assigned_shift

                # ✅ New logic to auto-fill shift_requested and weekdays_selected
                shift_type = assigned_shift.type.lower()
                if "weekendnight" in shift_type:
                    student_profile.shift_requested = "Weekend"
                    student_profile.weekdays_selected = {
                        "Friday": 4,
                        "Saturday": 5,
                        "Sunday": 6
                    }
                    weekdays_selected = [4, 5, 6]
                elif "weekendday" in shift_type:
                    student_profile.shift_requested = "Weekend"
                    student_profile.weekdays_selected = {
                        "Saturday": 5,
                        "Sunday": 6
                    }
                    weekdays_selected = [5, 6]
                elif "night" in shift_type:
                    student_profile.shift_requested = "Weekday"
                    student_profile.weekdays_selected = {
                        "Monday": 0,
                        "Tuesday": 1,
                        "Wednesday": 2,
                        "Thursday": 3,
                    }
                    weekdays_selected = [0, 1, 2, 3]
                else:
                    student_profile.shift_requested = "Weekday"
                    student_profile.weekdays_selected = {
                        "Monday": 0,
                        "Tuesday": 1,
                        "Wednesday": 2,
                        "Thursday": 3,
                        "Friday": 4
                    }
                    weekdays_selected = [0, 1, 2, 3, 4]

                student_profile.save()
                # Log activity
                
            except Shift.DoesNotExist:
                messages.error(request, "The specified shift does not exist.")
                return render(request, 'create_profile.html', {'form': profile_form, 'shifts': shifts})

            if student_profile.hours_requested:
                requested_hours = student_profile.hours_requested
                start_date = student_profile.start_date
                start_time = assigned_shift.start_time
                end_time = assigned_shift.end_time

                # shift_total_time = (datetime.combine(date.min, end_time) - datetime.combine(date.min, start_time)).seconds / 3600
                shift_total_time = calculate_shift_total_time(assigned_shift.start_time,assigned_shift.end_time)
                days_required = int(requested_hours // shift_total_time) + (1 if requested_hours % shift_total_time != 0 else 0)

                current_date = start_date
                working_days = 0
                while working_days < days_required:
                    if current_date.weekday() in weekdays_selected:
                        if validate_shift_capacity(request.user, current_date, assigned_shift):
                            VolunteerLog.objects.create(
                                student=student_profile,
                                date=current_date,
                                shift=assigned_shift,
                                start_time=start_time,
                                end_time=end_time,
                                hours_worked=shift_total_time,
                                status='Present'
                            )
                            working_days += 1
                        else:
                            messages.error(
                                request,
                                f"Capacity exceeded on {current_date.strftime('%Y-%m-%d')} for {assigned_shift.type} shift."
                            )
                            return render(request, 'create_profile.html', {'form': profile_form, 'shifts': shifts})
                    current_date += timedelta(days=1)

                student_profile.end_date = current_date - timedelta(days=1)
                student_profile.save()
                log_activity(
                    action_type="profile_created",
                    profile=student_profile,
                    created_by=request.user if request.user.is_authenticated else None,
                    shift_type=assigned_shift.type if assigned_shift else None,
                    shift_capacity=assigned_shift.max_students if assigned_shift else None,
                    start_date=student_profile.start_date,
                    end_date=student_profile.end_date,
                    extra_data={
                        "school": student_profile.school,
                        "email": student_profile.email,
                    }
                )

            threading.Thread(target=send_student_creation_email, args=(student_profile,)).start()

            messages.success(request, "Thank you for your submission! Please check your inbox and spam folder for a welcome email.")
            return redirect('student_profile_list')
        else:
            messages.error(request, "There were errors in the form. Please correct them.")
    else:
        profile_form = StudentProfileForm()

    # Only show non-hidden orientation dates in the form dropdown
    from .models import OrientationDate
    profile_form.fields['lchaim_orientation_date'].queryset = OrientationDate.objects.filter(hidden=False)

    return render(request, 'create_profile.html', {'form': profile_form, 'shifts': shifts})



@csrf_exempt  
def send_email(request):
    if request.method == 'POST':
        student_email = request.POST.get('email')
        
        if not student_email:
            return JsonResponse({'status': 'error', 'message': 'Student email is missing'})
        
        try:
            student = StudentProfile.objects.get(email=student_email)
            
            threading.Thread(target=send_student_creation_email, args=(student,)).start()
            
            return JsonResponse({'status': 'success'})
        
        except StudentProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Student not found'})
        except Exception as e:
            print(f"Error: {str(e)}")
            return JsonResponse({'status': 'error', 'message': 'An unexpected error occurred'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method. Only POST is allowed.'})
                         
                     
def send_student_creation_email(student):
    orientation_date = student.lchaim_orientation_date.date.strftime('%Y-%m-%d') if student.lchaim_orientation_date else student.start_date.strftime('%Y-%m-%d')
    orientation_description = student.lchaim_orientation_date.description if student.lchaim_orientation_date else None
    subject_q = 'New Student Profile Created'
    message = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                background-color: #f4f4f4;
                color: #333;
                padding: 20px;
            }}
            h3 {{
                color: #2c3e50;
            }}
            p {{
                font-size: 16px;
                line-height: 1.6;
            }}
            ul {{
                list-style-type: none;
                padding-left: 0;
            }}
            li {{
                margin-bottom: 8px;
            }}
            b {{
                color: #2980b9;
            }}
            .email-container {{
                background-color: #ffffff;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
        </style>
    </head>
    <body>
    <div class="email-container">
        <p>Dear {student.first_name},</p>

        <p>Welcome to L’Chaim! We are thrilled to have you join us and look forward to supporting you throughout your placement.</p>

        <p>Your orientation has been scheduled for <strong>{orientation_date} at 10:00 AM</strong>, and will take place at <strong>L’Chaim Retirement Home</strong>.</p>

        <p>During this session, we will walk you through important information to help ensure a smooth and successful start.</p>
        <p>Students may only begin their placement after the agreement has been signed.</p>

        <strong>Orientation Date & Time:</strong> {orientation_date} at 10:00 AM<br>
        <strong>Placement Location:</strong> L’Chaim Retirement Home – 718 Sheppard Ave West, Toronto, Ontario</p>

        <p>Before your orientation, please take some time to review the attached training documents, which outline our key policies and guidelines. Once you’ve read and understood them, kindly confirm your acknowledgement by replying to this email.</p>

        
<p style="padding-bottom:10px;">To get started with your placement, make sure to fill up the linked form below.</p>
        <p style="padding-bottom:10px;">👉 <strong>Student Placement Agreement Link:</strong> <a href="https://ca.services.docusign.net/webforms-ux/v1.0/forms/ae58e83ba452d7d8605aaf046d9e9f27" target="_blank">Click here.</a></p>


        <p>If you have any questions or need further clarification, don’t hesitate to reach out—we’re here to help.</p>

        <p>Wishing you a meaningful and enriching experience with us at L’Chaim!</p>

        <p>Warm regards,</p>

        <p><strong>Esther Davidov</strong><br>
        Students Coordinator</p>
    </div>
</body>

    </html>
    """

    volunteer_logs = VolunteerLog.objects.filter(student=student)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule for {student.first_name} {student.last_name}"

    headers = ['Date', 'Start Time', 'End Time', 'Shift']
    ws.append(headers)

    for log in volunteer_logs:
        start_time = log.start_time.strftime('%H:%M') if log.start_time else 'N/A'
        end_time = log.end_time.strftime('%H:%M') if log.end_time else 'N/A'
        shift_type = log.shift.type if log.shift and log.shift.type else 'N/A'
        
        ws.append([log.date, start_time, end_time, shift_type])

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    excel_filename = f"{student.first_name}_{student.last_name}_schedule.xlsx"

    email = EmailMessage(
    subject=subject_q,
    body=message,
    from_email=settings.DEFAULT_FROM_EMAIL,
    to=[
        student.email,
        student.college_contact_person_email
    ] if student.college_contact_person_email else [student.email]
)

    email.content_subtype = "html"

    email.attach(excel_filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    attachments_folder = os.path.join(settings.BASE_DIR, 'attachments')
    if os.path.exists(attachments_folder):
        for file_name in os.listdir(attachments_folder):
            if file_name.endswith('.pdf'):
                file_path = os.path.join(attachments_folder, file_name)
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as file:
                        email.attach(file_name, file.read(), 'application/pdf')
                else:
                    print(f"File {file_name} not found in attachments folder.")
    else:
        print(f"Attachments folder not found at {attachments_folder}.")
    email.send()



@login_required
def student_profile_list(request):
    search_query = request.GET.get("search", "")
    query = Q()  # empty query

    if search_query:
        query |= Q(first_name__icontains=search_query)
        query |= Q(last_name__icontains=search_query)
        query |= Q(email__icontains=search_query)



    # Sorting functionality
    sort_field = request.GET.get("sort_field", "").strip()
    sort_order = request.GET.get("sort_order", "asc").strip()

    if not sort_field:
        sort_field = "-id"

    valid_fields = {
        "first_name", "last_name", "email", "school", "start_date", "end_date", "status", "id", "lchaim_orientation_date"
    }

    if sort_field.lstrip("-") not in valid_fields:
        sort_field = "-id"

    if sort_field == "school":
        sort_field = "school" if sort_order == "asc" else "-school"
    if sort_field == "lchaim_orientation_date":
        sort_field = "lchaim_orientation_date" if sort_order == "asc" else "-lchaim_orientation_date"

    # Get distinct schools for dropdown
    schools = StudentProfile.objects.values_list("school", flat=True).distinct()

    # Get distinct orientation dates for dropdown
    orientation_dates = StudentProfile.objects.values("lchaim_orientation_date", "lchaim_orientation_date__date").distinct()


    print(orientation_dates)

    # School filtering
    school_filter = request.GET.get("school_filter", "").strip()
    if school_filter:
        query &= Q(school=school_filter)

    # Orientation date filtering
    orientation_date_filter = request.GET.get("orientation_date_filter", "").strip()
    if orientation_date_filter:
        query &= Q(lchaim_orientation_date=orientation_date_filter)

    # Fetch student profiles based on filters
    student_profiles = StudentProfile.objects.filter(query).order_by(sort_field)

    if request.GET.get("export") == "excel":
        return export_student_profiles_excel(student_profiles)
    if request.GET.get("export") == "pdf":
        return export_student_profiles_pdf(student_profiles)

    # Process weekdays_selected field
    for profile in student_profiles:
        raw_data = profile.weekdays_selected
        if isinstance(raw_data, str) and raw_data.strip():
            try:
                weekdays_dict = json.loads(raw_data.replace("'", '"'))
                if isinstance(weekdays_dict, dict):
                    profile.weekdays_display = ", ".join(weekdays_dict.keys())
                else:
                    profile.weekdays_display = "Invalid data"
            except json.JSONDecodeError:
                profile.weekdays_display = "Invalid data"
        else:
            profile.weekdays_display = "No weekdays selected"

    # Pagination
    paginator = Paginator(student_profiles, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # User permissions
    is_school_group = request.user.groups.filter(name='School').exists()
    is_admin = request.user.groups.filter(name='Admin').exists()

    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "sort_field": request.GET.get("sort_field", ""),
        "sort_order": request.GET.get("sort_order", ""),
        "schools": schools,  # List of schools for dropdown
        "school_filter": school_filter,  # Selected school filter
        "orientation_dates": orientation_dates,  # List of orientation dates for dropdown
        "orientation_date_filter": orientation_date_filter,  # Selected orientation date filter
        'is_school_group': is_school_group,
        'is_admin': is_admin
    }
    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']
    query_string = get_params.urlencode()
    context["query_string"] = query_string
    return render(request, "student_profile_list.html", context)

from django.template.loader import get_template
from xhtml2pdf import pisa
def export_student_profiles_pdf(student_profiles):
    # Process weekdays_display like in your main view
    for profile in student_profiles:
        raw_data = profile.weekdays_selected
        if isinstance(raw_data, str) and raw_data.strip():
            try:
                weekdays_dict = json.loads(raw_data.replace("'", '"'))
                profile.weekdays_display = ", ".join(weekdays_dict.keys()) if isinstance(weekdays_dict, dict) else "Invalid data"
            except json.JSONDecodeError:
                profile.weekdays_display = "Invalid data"
        else:
            profile.weekdays_display = "No weekdays selected"

    timestamp = datetime.now().strftime('%b-%d-%Y_%I-%M-%p')
    filename = f"student_profiles_{timestamp}.pdf"

    response = render_to_pdf("student_profiles_pdf.html", {"student_profiles": student_profiles})
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
def export_student_profiles_excel(student_profiles):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Student Profiles"

    # Define column headers
    columns = [
        "First Name", "Last Name", "Email", "School", "L'Chaim Orientation Date",
        "Start Date", "End Date", "Status", "Weekdays Selected"
    ]
    worksheet.append(columns)

    # Styling for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        worksheet.column_dimensions[get_column_letter(col_num)].width = 22

    # Add data rows
    for row_num, profile in enumerate(student_profiles, start=2):
        raw_data = profile.weekdays_selected
        if isinstance(raw_data, str) and raw_data.strip():
            try:
                weekdays_dict = json.loads(raw_data.replace("'", '"'))
                weekdays_display = ", ".join(weekdays_dict.keys()) if isinstance(weekdays_dict, dict) else "Invalid data"
            except json.JSONDecodeError:
                weekdays_display = "Invalid data"
        else:
            weekdays_display = "No weekdays selected"

        row_data = [
            profile.first_name,
            profile.last_name,
            profile.email,
            profile.school,
            profile.lchaim_orientation_date.date.strftime('%Y-%m-%d') if profile.lchaim_orientation_date else "",
            profile.start_date.strftime('%Y-%m-%d') if profile.start_date else "",
            profile.end_date.strftime('%Y-%m-%d') if profile.end_date else "",
            profile.status,
            weekdays_display
        ]

        # Add each cell with styling
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = center_align

            # Add alternate row coloring
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Freeze header row
    worksheet.freeze_panes = "A2"

    # Add auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    timestamp = datetime.now().strftime('%b-%d-%Y_%I-%M-%p')
    filename = f"student_profiles_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)
    return response


@login_required
def student_graduated_list(request):
    query = request.GET.get('q', '')
    orientation_date = request.GET.get('orientation_date', '')

    profiles = StudentProfile.objects.filter(status='Graduated')


    if query:
        profiles = profiles.filter(
            first_name__icontains=query
        ) | profiles.filter(
            last_name__icontains=query
        ) | profiles.filter(
            email__icontains=query
        )

    if orientation_date:
        profiles = profiles.filter(lchaim_orientation_date=orientation_date)

    paginator = Paginator(profiles, 10) 
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

  
    is_school_group = request.user.groups.filter(name='School').exists()

    return render(request, 'graduated.html', {
        'page_obj': page_obj,
        'query': query,
        'orientation_date': orientation_date,
        'is_school_group': is_school_group,  
    })

@login_required
def reports_list(request):
    shift_filter = request.GET.get('shift')
    date_filter = request.GET.get('start_date')
    export_format = request.GET.get('export')
    log_date_filter = request.GET.get('log_date')

    # students = StudentProfile.objects.select_related('assigned_shift').all()

    # if shift_filter:
    #     students = students.filter(assigned_shift__type=shift_filter)

    # # if date_filter:
    # #     try:
    # #         date_obj = datetime.strptime(date_filter, "%Y-%m-%d").date()
    # #         students = students.filter(start_date=date_obj)
    # #     except ValueError:
    # #         pass

    # if log_date_filter:
    #     try:
    #         log_date = datetime.strptime(log_date_filter, "%Y-%m-%d").date()
    #         students = students.filter(volunteer_logs__date=log_date).distinct()
    #     except ValueError:
    #         pass

    # paginator = Paginator(students, 10)
    # page_number = request.GET.get("page")
    # page_obj = paginator.get_page(page_number)
    
    
    logs = VolunteerLog.objects.all()
    
    # Apply filters if present
    if shift_filter:
        logs = logs.filter(shift__type=shift_filter)
    elif not shift_filter:
    # Exclude logs with no shift assigned (null)
        logs = logs.exclude(shift__isnull=True)
    
    if log_date_filter:
        try:
            log_date = datetime.strptime(log_date_filter, "%Y-%m-%d").date()
            logs = logs.filter(date=log_date)
        except ValueError:
            log_date = None
    else:
        log_date = None
        
     # Extract related student profiles
    student_ids = logs.values_list('student_id', flat=True).distinct()
    students_qs = StudentProfile.objects.filter(id__in=student_ids)
    
    
    paginator = Paginator(students_qs.order_by('first_name', 'last_name'), 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    

    if export_format == "pdf":
        return export_pdf(students_qs)
    elif export_format == "excel":
        return export_excel(students_qs)

    shifts = Shift.objects.all()

    # Get unique start dates manually (sorted)
    start_dates = (
        StudentProfile.objects.exclude(start_date__isnull=True)
        .order_by("start_date")
        .values_list("start_date", flat=True)
        .distinct()
    )

    log_dates = (
        VolunteerLog.objects.exclude(date__isnull=True)
        .order_by("date")
        .values_list("date", flat=True)
        .distinct()
    )

    return render(request, "reports.html", {
        "students": page_obj,
        "shifts": shifts,
        "shift_filter": shift_filter,
        "date_filter": date_filter,
        "log_date_filter": log_date_filter,
        "start_dates": start_dates,
        "log_dates": log_dates,
    })


def export_pdf(students_page):
    now = datetime.now()
    template = get_template("report_pdf.html")
    
    # Pass 'now' to the template for timestamp rendering
    html = template.render({
        "students": students_page,
        "now": now
    })

    # Format the filename with timestamp
    timestamp_str = now.strftime("%Y-%m-%d_%H-%M")
    filename = f"students_report_{timestamp_str}.pdf"

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    pisa.CreatePDF(html, dest=response)
    return response

def export_excel(students_page):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Reports"

    # Define header and styling
    headers = ["First Name", "Last Name", "Shift", "Start Date"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Add headers with styling
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Add student data
    for row_num, student in enumerate(students_page, start=2):
        row_data = [
            student.first_name,
            student.last_name,
            student.assigned_shift.type if student.assigned_shift else "Not Assigned",
            student.start_date.strftime("%Y-%m-%d") if student.start_date else ""
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left")
            cell.border = thin_border

    # Set column widths based on max content
    column_widths = [15, 15, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Prepare response
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"students_report_{timestamp}.xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@admin_required
def update_student_profile(request, pk):
    profile = get_object_or_404(StudentProfile, pk=pk)

    if request.method == "POST":

        form = StudentProfileForm(request.POST, request.FILES, instance=profile)

        if form.is_valid():
            student_profile = form.save()

            if 'documents' in request.FILES:
                files = request.FILES.getlist('documents') 
                for file in files:
                    StudentFile.objects.create(student=student_profile, file=file)

            return redirect('student_profile_list')
        else :
            print("File Error:", form.errors)  

    else:
        form = StudentProfileForm(instance=profile)

    existing_files = StudentFile.objects.filter(student=profile)

    return render(request, 'update_profile.html', {
        'form': form,
        'existing_files': existing_files
    })


@admin_required
def delete_student_profile(request, pk):
    profile = get_object_or_404(StudentProfile, pk=pk)
    if request.method == "POST":
        profile.delete()
    return redirect('student_profile_list')

@admin_required
def log_volunteer_hours(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    if request.method == "POST":
        form = VolunteerLogForm(request.POST)
        if form.is_valid():
            volunteer_log = form.save(commit=False)
            volunteer_log.student = student
            volunteer_log.save()
            
            student.hours_completed += volunteer_log.hours_worked
            student.save()
            return redirect('student_profile_list')
    else:
        form = VolunteerLogForm()
    return render(request, 'log_volunteer_hours.html', {'form': form, 'student': student})

@admin_required
def student_attendance(request):
    date_str = request.GET.get('date') or timezone.now().strftime('%Y-%m-%d')
    shift_filter = request.GET.get('shift', '').strip()
    try:
        selected_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = timezone.now().date()

    print(f"[DEBUG] Selected date: {selected_date}")
    print(f"[DEBUG] Shift filter: {shift_filter}")

    # Debug: Print logs with no shift or shift type null/empty for this date
    logs_no_shift = VolunteerLog.objects.filter(date=selected_date, shift__isnull=True)
    if logs_no_shift.exists():
        print("[DEBUG] (Attendance) VolunteerLog entries with shift=None:")
        for log in logs_no_shift:
            print(f"  id={log.id}, student={log.student}, status={log.status}")

    logs_null_type = VolunteerLog.objects.filter(date=selected_date, shift__type__isnull=True)
    if logs_null_type.exists():
        print("[DEBUG] (Attendance) VolunteerLog entries with shift.type=None:")
        for log in logs_null_type:
            print(f"  id={log.id}, student={log.student}, shift_id={log.shift_id}, status={log.status}")

    logs_empty_type = VolunteerLog.objects.filter(date=selected_date, shift__type='')
    if logs_empty_type.exists():
        print("[DEBUG] (Attendance) VolunteerLog entries with shift.type='':")
        for log in logs_empty_type:
            print(f"  id={log.id}, student={log.student}, shift_id={log.shift_id}, status={log.status}")

    # Fetch logs for the selected date
    logs = VolunteerLog.objects.filter(date=selected_date)
    print(f"[DEBUG] VolunteerLog entries for date {selected_date}: {logs.count()}")
    if shift_filter:
        logs = logs.filter(shift__type=shift_filter)
        print(f"[DEBUG] VolunteerLog entries after shift filter '{shift_filter}': {logs.count()}")

    # Print first few logs for inspection
    debug_logs = list(logs.select_related('student', 'shift')[:5])
    for log in debug_logs:
        print(f"[DEBUG] Log: student={log.student.first_name} {log.student.last_name}, status={log.status}, shift={log.shift.type if log.shift else None}")

    # Prepare student_logs for the template
    student_logs = []
    present_count = 0
    absent_count = 0
    total_count = logs.count()

    for log in logs.select_related('student', 'shift'):
        if log.status == 'Present':
            present_count += 1
        elif log.status == 'Absent':
            absent_count += 1
        # Format hours_worked for display
        if log.hours_worked is not None:
            hours = int(log.hours_worked)
            minutes = round((log.hours_worked - hours) * 60)
            log.hours_worked_display = f"{hours}:{minutes:02}"
        else:
            log.hours_worked_display = "00:00"
        student_logs.append({'student': log.student, 'log': log})

    from .models import Shift, StudentProfile
    all_shifts = Shift.objects.all()  # Always pass all shifts from backend
    all_students = StudentProfile.objects.all()

    context = {
        'students': student_logs,
        'selected_date': selected_date,
        'all_shifts': all_shifts,  # Use this for filter and modal
        'shift_filter': shift_filter,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_count': total_count,
        'all_students': all_students,  # For modal dropdown
    }

    return render(request, 'attendance/attendance.html', context)

def update_attendance(request, student_id, date):
    student = get_object_or_404(StudentProfile, id=student_id)
    today = timezone.now().date()
    log, created = VolunteerLog.objects.get_or_create(student=student, date=date)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        notes = request.POST.get('notes')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        hours_worked = request.POST.get('hours_worked')

        if start_time and end_time:
            start_time = datetime.strptime(start_time, '%H:%M').time()
            end_time = datetime.strptime(end_time, '%H:%M').time() 
            log.start_time = start_time
            log.end_time = end_time

        log.status = status
        log.notes = notes

        if hours_worked:
            if isinstance(hours_worked, str) and ':' in hours_worked:
                hours, minutes = map(int, hours_worked.split(':'))
                total_hours = hours + minutes / 60.0  
                log.hours_worked = total_hours
            else:
                log.hours_worked = float(hours_worked)
        
        log.save()

        return JsonResponse({'success': True})
    return JsonResponse({'success': False})


def student_details(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    
    volunteer_logs = VolunteerLog.objects.filter(student=student).order_by('date')  # Ascending order
    student_files = StudentFile.objects.filter(student=student)
    
    filename_pattern = re.compile(r'([^/]+)$')
    
    # Process the file names using regex
    for file in student_files:
        match = filename_pattern.search(file.file.name)
        if match:
            file.file = match.group(0)  
    
    total_hours = 0
    for log in volunteer_logs:
        if log.status == 'Present':
            total_hours += log.hours_worked if log.hours_worked else 7
    
    if request.method == 'POST':
        log_id = request.POST.get('log_id')
        action = request.POST.get('action')
        
        # Handle toggle action to mark log as absent
        if action == 'toggle':
            log = get_object_or_404(VolunteerLog, id=log_id, student=student)
            previous_log_data = {
                'start_time': log.start_time,
                'end_time': log.end_time,
                'hours_worked': log.hours_worked,
                'notes': log.notes
            }
            
            log.status = 'Absent'
            log.hours_worked = 0
            log.start_time = None
            log.end_time = None
            log.notes = "Marked as Absent"
            log.save()
            
            last_log = VolunteerLog.objects.filter(student=student).order_by('-date').first()
            if last_log:
                new_date = last_log.date + timedelta(days=1)  # Set the date as the last date + 1
                
                # Create a new log entry with the new date and copy data from the previous log
                new_log = VolunteerLog.objects.create(
                    student=student,
                    date=new_date,
                    status='Present',
                    start_time=previous_log_data['start_time'],
                    end_time=previous_log_data['end_time'],
                    hours_worked=previous_log_data['hours_worked'],
                    notes=previous_log_data['notes']
                )

            return redirect('student_details', student_id=student.id)

        # Handle delete action
        if action == 'delete':
            log = get_object_or_404(VolunteerLog, id=log_id, student=student)
            log.delete()
            return redirect('student_details', student_id=student.id)
    
    return render(request, 'student_details.html', {
        'student': student,
        'volunteer_logs': volunteer_logs,
        'total_hours': total_hours,
        'student_files': student_files
    })


@admin_required
def student_logs(request):
    query = request.GET.get('q', '').strip()
    
    if query:
        students = StudentProfile.objects.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query)
        )
    else:
        students = StudentProfile.objects.all()

    students = students.order_by(
        Case(
            When(status='Training', then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        ),
        'first_name'  
    )

    student_data = []
    for student in students:
        today = now().date()
        volunteer_logs = VolunteerLog.objects.filter(student=student, date__lte=today).order_by('date')

        total_hours = 0
        latest_date = None

        for log in volunteer_logs:
            if log.status == 'Present':
                total_hours += log.hours_worked if log.hours_worked else 8
            latest_date = log.date if not latest_date or log.date > latest_date else latest_date
        
        total_hours = max(0, total_hours) 
        
        student.hours_completed = total_hours
        student.save()

        hours_requested = student.hours_requested if hasattr(student, 'hours_requested') else 0
        remaining_hours = hours_requested - total_hours

        progress_width = (total_hours / hours_requested) * 100 if hours_requested > 0 else 0
        progress_width = round(progress_width, 2)  

        if student.status == 'Graduated':
            wfstatus = 'Inactive'  
        else:
            wfstatus = 'Active'  

        student_data.append({
            'student': student,
            'volunteer_logs': volunteer_logs,
            'total_hours': total_hours,
            'hours_requested': student.hours_requested,
            'remaining_hours': remaining_hours,
            'progress_width': progress_width,
            'latest_date': latest_date,
            'status': wfstatus  
        })

    return render(request, 'student_logs.html', {'student_data': student_data, 'query': query})

def get_start_of_week(date):
    """Get the start of the week (Monday)."""
    start = date - timedelta(days=date.weekday())
    return start.replace(hour=0, minute=0, second=0, microsecond=0)

def get_end_of_week(date):
    """Get the end of the week (Sunday)."""
    end = date + timedelta(days=(6 - date.weekday()))
    return end.replace(hour=23, minute=59, second=59, microsecond=0)


from django.utils.timezone import localtime
@admin_required
def calendar_student_logs(request):
    current_date = timezone.now()
    week_start_date = get_start_of_week(current_date)
    week_end_date = get_end_of_week(current_date)

    # Debug: Print logs with no shift or shift type null/empty for this week
    logs_no_shift = VolunteerLog.objects.filter(date__gte=week_start_date, date__lte=week_end_date, shift__isnull=True)
    if logs_no_shift.exists():
        print("[DEBUG] (Calendar) VolunteerLog entries with shift=None:")
        for log in logs_no_shift:
            print(f"  id={log.id}, student={log.student}, date={log.date}, status={log.status}")

    logs_null_type = VolunteerLog.objects.filter(date__gte=week_start_date, date__lte=week_end_date, shift__type__isnull=True)
    if logs_null_type.exists():
        print("[DEBUG] (Calendar) VolunteerLog entries with shift.type=None:")
        for log in logs_null_type:
            print(f"  id={log.id}, student={log.student}, date={log.date}, shift_id={log.shift_id}, status={log.status}")

    logs_empty_type = VolunteerLog.objects.filter(date__gte=week_start_date, date__lte=week_end_date, shift__type='')
    if logs_empty_type.exists():
        print("[DEBUG] (Calendar) VolunteerLog entries with shift.type='':")
        for log in logs_empty_type:
            print(f"  id={log.id}, student={log.student}, date={log.date}, shift_id={log.shift_id}, status={log.status}")

    students = StudentProfile.objects.all()
    student_data = []

    for student in students:
        volunteer_logs = VolunteerLog.objects.filter(student=student, status='Present').order_by('start_time')

        total_hours = 0
        formatted_logs = []

        for log in volunteer_logs:
            if log.hours_worked is None:
                total_hours += 7
            else:
                total_hours += log.hours_worked

            formatted_logs.append({
                'date': log.date.isoformat(),
                'status': log.status,
                'start_time': log.start_time.strftime('%H:%M') if log.start_time else None,
                'end_time': log.end_time.strftime('%H:%M') if log.end_time else None,
                'hours_worked': float(log.hours_worked) if isinstance(log.hours_worked, decimal.Decimal) else log.hours_worked,
                'notes': log.notes or '',
                'shift': {
                    'type': log.shift.type if log.shift else None,
                    'start_time': log.shift.start_time.strftime('%H:%M') if log.shift and log.shift.start_time else None,
                    'end_time': log.shift.end_time.strftime('%H:%M') if log.shift and log.shift.end_time else None,
                } if log.shift else None,
            })

        student_data.append({
            'id': student.id,
            'phone': student.phone,
            'name': f"{student.first_name} {student.last_name}",
            'logs': formatted_logs,
            'total_hours': float(total_hours) if isinstance(total_hours, decimal.Decimal) else total_hours
        })
    student_data_json = json.dumps(student_data)
    return render(request, 'calendar.html', {
        'student_data': student_data_json,
        'week_start_date': week_start_date,
        'week_end_date': week_end_date,
        'current_date':localtime().isoformat() 
    })

def download_excel(request, selected_date):
    date = datetime.strptime(selected_date, '%Y-%m-%d').date()

    student_data = []

    students = StudentProfile.objects.all()
    for student in students:
        volunteer_logs = VolunteerLog.objects.filter(student=student, date=date)

        for log in volunteer_logs:
            student_data.append({
                'Student Name': f"{student.first_name} {student.last_name}",
                'School': student.school,
                'Date of Attendance': log.date,
                'Hours Worked': log.hours_worked if log.hours_worked is not None else 'N/A',
                'Status': log.status,
                'Notes': log.notes if log.notes else 'N/A',
            })

    if not student_data:
        return HttpResponse("No attendance data found for the selected date.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance_{date}"

    headers = ['Student Name', 'School', 'Date of Attendance', 'Hours Worked', 'Status', 'Notes']
    ws.append(headers)

    for entry in student_data:
        ws.append([entry['Student Name'], entry['School'], entry['Date of Attendance'], entry['Hours Worked'], entry['Status'], entry['Notes']])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_{date}.xlsx'

    wb.save(response)
    return response



def download_excel_user_wise(request, id):
    student_data = []
    students = StudentProfile.objects.filter(id=id)
    
    for student in students:
        total_hours = 0  
        
        volunteer_logs = VolunteerLog.objects.filter(student=student, status='Present')
        
        for log in volunteer_logs:
            total_hours += log.hours_worked if log.hours_worked else 7

        hours_requested = student.hours_requested if hasattr(student, 'hours_requested') else 0
        remaining_hours = hours_requested - total_hours
        
        for log in volunteer_logs:
            student_data.append({
                'Student Name': f"{student.first_name} {student.last_name}",
                'Email': student.email,
                'Phone': student.phone,
                'School': student.school,
                'Date of Attendance': log.date,
                'Start Time': log.start_time,
                'End Time': log.end_time,
                'Attendance': log.status,
                'Notes': log.notes if log.notes else 'N/A',
                'Completed Hours': log.hours_worked if log.hours_worked is not None else 'N/A',
                'Remaining Hours': remaining_hours,
                'Status': student.status,
            })

    if not student_data:
        return HttpResponse("No attendance data found for the selected student.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ['Student Name', 'Email', 'Phone', 'School', 'Date of Attendance', 'Start Time', 'End Time', 'Attendance', 'Notes', 'Completed Hours', 'Remaining Hours', 'Status']
    ws.append(headers)

    for entry in student_data:
        ws.append([
            entry['Student Name'],
            entry['Email'],
            entry['Phone'],
            entry['School'],
            entry['Date of Attendance'],
            entry['Start Time'],
            entry['End Time'],
            entry['Attendance'],
            entry['Notes'],
            entry['Completed Hours'],
            entry['Remaining Hours'],
            entry['Status'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={student.first_name}_{student.last_name}_report.xlsx'

    wb.save(response)
    return response

@admin_required
def create_college(request):
    if request.method == 'POST':
        form = CollegeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_profile_list')  
    else:
        form = CollegeForm()

    return render(request, 'create_college.html', {'form': form})

@admin_required
def student_file_details(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    file_logs = StudentFile.objects.filter(student=student)  
    
    return render(request, 'studentfile.html', {
        'student': student,
        'file_logs': file_logs,  
    })

def download_file(request, file_id):
    try:
        file_log = StudentFile.objects.get(id=file_id)
        file_path = file_log.file.path 

        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        response['Content-Disposition'] = f'attachment; filename="{file_log.file.name}"'
        return response
    except StudentFile.DoesNotExist:
        raise Http404("File not found.")


def remove_file(request, pk):
    student_file = get_object_or_404(StudentFile, pk=pk)
    student_file.delete()
    return redirect('update_student_profile', pk=student_file.student.pk)

class MarkGraduateAPIView(APIView):
    def patch(self, request, pk):
        student = get_object_or_404(StudentProfile, pk=pk)

        if student.status == 'Graduated':
            return Response({"detail": "Student is already graduated."}, status=status.HTTP_400_BAD_REQUEST)

        student.status = 'Graduated'
        student.save()

        return Response({
            "detail": f"Student {student.first_name} {student.last_name} has been marked as graduated.",
            "id": student.id,
            "status": student.status
        }, status=status.HTTP_200_OK)
    
def orientation_date_list(request):
    is_admin = request.user.groups.filter(name='Admin').exists()
    if is_admin:
        # Show all dates to admin, both hidden and not hidden
        dates = OrientationDate.objects.all()
    else:
        # Only show dates that are not hidden to non-admins
        dates = OrientationDate.objects.filter(hidden=False)
    return render(request, 'orientation_date_list.html', {'dates': dates, 'is_admin': is_admin})


# Add a new orientation date
@admin_required
def add_orientation_date(request):
    if request.method == 'POST':
        form = OrientationDateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('orientation_date_list')
    else:
        form = OrientationDateForm()
    return render(request, 'add_orientation_date.html', {'form': form})

# Edit an existing orientation date

@admin_required
def edit_orientation_date(request, pk):
    orientation_date = get_object_or_404(OrientationDate, pk=pk)
    if request.method == 'POST':
        form = OrientationDateForm(request.POST, instance=orientation_date)
        if form.is_valid():
            form.save()
            return redirect('orientation_date_list')
    else:
        form = OrientationDateForm(instance=orientation_date)
    return render(request, 'edit_orientation_date.html', {'form': form})

# Delete an orientation date
@admin_required
def delete_orientation_date(request, pk):
    orientation_date = get_object_or_404(OrientationDate, pk=pk)
    if request.method == 'POST':
        orientation_date.delete()
        return redirect('orientation_date_list')
    return HttpResponse(status=405)  # Method not allowed

def delete_log(request, log_id):
    log = get_object_or_404(VolunteerLog, id=log_id)
    
    # Delete the log
    log.delete()
    
    # Optionally, add a success message
    messages.success(request, "Volunteer log has been deleted successfully.")
    
    # Redirect back to the student's details page or wherever necessary
    return redirect('student_details', student_id=log.student.id)

def student_counts_by_date_and_shift(request):
    # Debug: Print logs with no shift or shift type null/empty
    logs_no_shift = VolunteerLog.objects.filter(shift__isnull=True)
    if logs_no_shift.exists():
        print("[DEBUG] VolunteerLog entries with shift=None:")
        for log in logs_no_shift:
            print(f"  id={log.id}, student={log.student}, date={log.date}, status={log.status}")

    logs_null_type = VolunteerLog.objects.filter(shift__type__isnull=True)
    if logs_null_type.exists():
        print("[DEBUG] VolunteerLog entries with shift.type=None:")
        for log in logs_null_type:
            print(f"  id={log.id}, student={log.student}, date={log.date}, shift_id={log.shift_id}, status={log.status}")

    logs_empty_type = VolunteerLog.objects.filter(shift__type='')
    if logs_empty_type.exists():
        print("[DEBUG] VolunteerLog entries with shift.type='':")
        for log in logs_empty_type:
            print(f"  id={log.id}, student={log.student}, date={log.date}, shift_id={log.shift_id}, status={log.status}")

    # Filter only logs where status is Present to count actual attendance
    data = (
        VolunteerLog.objects
        .filter(status='Present')  
        .values('date', 'shift__type')
        .annotate(student_count=Count('student', distinct=True))
        .order_by('date', 'shift__type')
    )

    # Format response to match frontend expectations
    formatted = [
        {
            "start_date": entry["date"].isoformat(),
            "assigned_shift__type": entry["shift__type"],
            "student_count": entry["student_count"]
        }
        for entry in data
    ]

    return JsonResponse(formatted, safe=False)
    

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import StudentProfile

@require_POST
def update_start_date(request, pk):
    profile = get_object_or_404(StudentProfile, pk=pk)
    new_start_date = request.POST.get('new_start_date')
    if new_start_date:
        profile.start_date = new_start_date
        profile.save()
        messages.success(request, f"Start date updated for {profile.first_name} {profile.last_name}.")
    else:
        messages.error(request, "Invalid start date.")
    return redirect('student_profile_list')

@admin_required
@require_POST
def hide_orientation_date(request, pk):
    orientation_date = get_object_or_404(OrientationDate, pk=pk)
    orientation_date.hidden = True
    orientation_date.save()
    return redirect('orientation_date_list')

@admin_required
@require_POST
def unhide_orientation_date(request, pk):
    orientation_date = get_object_or_404(OrientationDate, pk=pk)
    orientation_date.hidden = False
    orientation_date.save()
    return redirect('orientation_date_list')

@csrf_exempt
@admin_required
def add_extended_volunteer_log(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            date_str = data.get('date')
            shift_id = data.get('shift_id')
            start_time = data.get('start_time')
            end_time = data.get('end_time')
            hours_worked = data.get('hours_worked')
            notes = data.get('notes', '')
            status = data.get('status', 'Present')

            student = StudentProfile.objects.get(id=student_id)
            shift = Shift.objects.get(id=shift_id)
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            start_time_obj = datetime.strptime(start_time, '%H:%M').time()
            end_time_obj = datetime.strptime(end_time, '%H:%M').time()
            hours_worked = float(hours_worked)

            # Prevent duplicate: check for any log (regular or extended) for this student/date/shift
            duplicate_log = VolunteerLog.objects.filter(student=student, date=date_obj, shift=shift).exists()
            if duplicate_log:
                return JsonResponse({'success': False, 'error': 'A log for this student, date, and shift already exists.'}, status=400)

            # Remove any existing extended log for this student/date/shift (shouldn't be needed now, but keep for safety)
            VolunteerLog.objects.filter(student=student, date=date_obj, shift=shift, extended=True).delete()

            log = VolunteerLog.objects.create(
                student=student,
                date=date_obj,
                shift=shift,
                start_time=start_time_obj,
                end_time=end_time_obj,
                hours_worked=hours_worked,
                notes=notes,
                status=status,
                extended=True
            )
            return JsonResponse({'success': True, 'log_id': log.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

# New: Delete volunteer log by ID (admin only, POST only)
@csrf_exempt
@admin_required
@require_POST
def delete_volunteer_log(request, log_id):
    try:
        log = VolunteerLog.objects.get(id=log_id)
        log.delete()
        return JsonResponse({'success': True})
    except VolunteerLog.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Log not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@admin_required
def admin_dashboard(request):
    from .models import ActivityLog, StudentProfile, Shift
    # Filtering, searching, sorting
    logs = ActivityLog.objects.select_related('profile', 'created_by')
    search = request.GET.get('search', '').strip()
    action_type = request.GET.get('action_type', '').strip()
    sort = request.GET.get('sort', '-created_at')

    if search:
        logs = logs.filter(
            Q(profile__first_name__icontains=search) |
            Q(profile__last_name__icontains=search) |
            Q(created_by__username__icontains=search) |
            Q(shift_type__icontains=search)
        )
    if action_type:
        logs = logs.filter(action_type=action_type)
    if sort:
        logs = logs.order_by(sort)
    else:
        logs = logs.order_by('-created_at')

    logs = logs[:200]  # Limit for performance
    total_logs = ActivityLog.objects.count()
    # Stats
    total_profiles = StudentProfile.objects.count()
    total_shifts = Shift.objects.count()
    total_admins = request.user.__class__.objects.filter(groups__name='Admin').count()
    profiles_created_today = StudentProfile.objects.filter(created_at__date=timezone.now().date()).count() if hasattr(StudentProfile, 'created_at') else None
    # For filter dropdown
    action_types = ActivityLog.ACTION_CHOICES
    context = {
        'logs': logs,
        'total_logs': total_logs,
        'total_profiles': total_profiles,
        'total_shifts': total_shifts,
        'total_admins': total_admins,
        'profiles_created_today': profiles_created_today,
        'search': search,
        'action_type': action_type,
        'sort': sort,
        'action_types': action_types,
    }
    return render(request, 'admin_dashboard.html', context)